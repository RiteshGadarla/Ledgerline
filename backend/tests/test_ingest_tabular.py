import io

import openpyxl  # type: ignore[import-untyped]

from ingest.tabular import parse_table, sniff_content_type
from money.result import Err, Ok


def _csv_bytes(text: str) -> bytes:
    return text.encode()


def test_plain_csv_parses() -> None:
    content = _csv_bytes("id,amount\nINV1,100.00\nINV2,200.00\n")
    result = parse_table(content, "invoices.csv")
    assert isinstance(result, Ok)
    assert result.value.headers == ["id", "amount"]
    assert len(result.value.rows) == 2
    assert result.value.rows[0] == {"id": "INV1", "amount": "100.00"}


def test_csv_with_bom() -> None:
    content = b"\xef\xbb\xbfid,amount\nINV1,100.00\n"
    result = parse_table(content, "invoices.csv")
    assert isinstance(result, Ok)
    assert result.value.headers == ["id", "amount"]


def test_csv_with_indian_grouped_amount_passes_through_as_text() -> None:
    content = _csv_bytes("id,amount\nINV1,\"1,23,456.78\"\n")
    result = parse_table(content, "invoices.csv")
    assert isinstance(result, Ok)
    assert result.value.rows[0]["amount"] == "1,23,456.78"


def test_csv_with_dd_mmm_yy_dates_passes_through_as_text() -> None:
    content = _csv_bytes("id,issued_at\nINV1,05-Jan-24\n")
    result = parse_table(content, "invoices.csv")
    assert isinstance(result, Ok)
    assert result.value.rows[0]["issued_at"] == "05-Jan-24"


def test_totals_row_is_excluded() -> None:
    content = _csv_bytes("id,amount\nINV1,100.00\nINV2,200.00\nTotal,300.00\n")
    result = parse_table(content, "invoices.csv")
    assert isinstance(result, Ok)
    assert len(result.value.rows) == 2
    assert all(row["id"] != "Total" for row in result.value.rows)


def test_empty_csv_is_an_error() -> None:
    result = parse_table(b"", "empty.csv")
    assert isinstance(result, Err)
    assert "empty" in result.reason


def test_image_renamed_to_csv_is_rejected() -> None:
    png_bytes = b"\x89PNG\r\n\x1a\n" + b"\x00" * 20
    result = parse_table(png_bytes, "fake.csv")
    assert isinstance(result, Err)
    assert "image" in result.reason


def test_xlsx_with_merged_header_cells() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["id", "amount"])
    sheet.merge_cells("A1:B1")  # a merged header banner above the real headers, common in exports
    sheet.append(["INV1", "100.00"])
    sheet.append(["INV2", "200.00"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    result = parse_table(buffer.getvalue(), "invoices.xlsx")
    assert isinstance(result, Ok)
    assert len(result.value.rows) == 2


def test_50k_row_csv_parses_without_error() -> None:
    lines = ["id,amount"] + [f"INV{i},{i}.00" for i in range(50_000)]
    content = "\n".join(lines).encode()
    result = parse_table(content, "big.csv")
    assert isinstance(result, Ok)
    assert len(result.value.rows) == 50_000


def test_sniff_content_type_detects_xlsx_by_magic_bytes() -> None:
    workbook = openpyxl.Workbook()
    buffer = io.BytesIO()
    workbook.save(buffer)
    assert sniff_content_type(buffer.getvalue()) == "xlsx"


def test_sniff_content_type_detects_pdf_by_magic_bytes() -> None:
    assert sniff_content_type(b"%PDF-1.4\n...") == "pdf"
